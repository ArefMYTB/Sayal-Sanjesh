import os
import random
import string
import openpyxl
from django.db.models import Sum
from datetime import datetime
from persiantools.jdatetime import JalaliDate
from persiantools import digits
from openpyxl import load_workbook
from Authorization.TokenManager import token_to_user_id
from SayalSanjesh.Serializers import wrong_token_result, status_success_result, wrong_data_result, wrong_result
from Authorization.Serializers.AdminsSerializer import AdminsSerializer
from SayalSanjesh.models import Bills, WaterMeters, WaterMetersConsumptions, Pattern, WaterMetersTags, \
    WaterMetersProjects
from Authorization.models.Admins import Admins
from MQQTReceiver.publisher import publish_message_to_client


class BillsSerializer:
    """
        This class is responsible for serializing data in our Django app
    """

    @staticmethod
    def admin_get_all_bills_serializer(token, page, count, water_meter_serial, bill_serial, bill_start_date, user_id,
                                       bill_end_date, payment_dead_line, bill_create_date, project_id,
                                       user_phone_number):
        """
            param : [token, page, count, water_meter_serial, bill_serial, bill_start_date, user_id,
                                           bill_end_date, payment_dead_line, bill_create_date, project_id,
                                           user_phone_number]

            return :
            A tuple containing a boolean indicating the success or failure of the operation, and a list of serialized data
            results.  it returns a false status along with an error message.
        """
        token_result = token_to_user_id(token)
        if token_result["status"] == "OK":
            admin_id = token_result["data"]["user_id"]
            if AdminsSerializer.admin_check_permission(admin_id, ['BillManaging', 'ViewDevice']):
                offset = int((page - 1) * count)
                limit = int(count)
                filters = {
                    "bill_water_meter__water_meter_serial": water_meter_serial,
                    "bill_serial": bill_serial,
                    "bill_water_meter__water_meter_user__user_id": user_id,
                    "bill_water_meter__water_meter_user__user_phone": user_phone_number,
                    "bill_water_meter__water_meter_project__water_meter_project_id": project_id,
                    "bill_start_date": bill_start_date,
                    "bill_end_date": bill_end_date,
                    "payment_dead_line": payment_dead_line,
                    "bill_create_date": bill_create_date
                }
                filters = {k: v for k, v in filters.items() if v is not None}
                try:
                    bills = Bills.objects.filter(**filters)
                except:
                    wrong_data_result["farsi_message"] = "داده های ورودی اشتباه است !!!"
                    wrong_data_result["english_message"] = "The input data is incorrect !!!"
                    return False, wrong_data_result
                bill_pagination = bills.order_by('bill_create_date')[offset:offset + limit]
                bill_results = []
                for bill in bill_pagination:
                    bill = bill.as_dict()
                    bill['all_bill_count'] = bills.count()
                    bill_results.append(bill)
                return True, bill_results
            else:
                return False, wrong_token_result
        else:

            return False, wrong_token_result

    @staticmethod
    def admin_get_one_bill_serializer(token, bill_id, bill_serial):
        """
            param : [token, bill_id, bill_serial]

            return :
            A tuple containing a boolean indicating the success or failure of the operation, and a list of serialized data
            results.  it returns a false status along with an error message.
        """
        token_result = token_to_user_id(token)
        if token_result["status"] == "OK":
            admin_id = token_result["data"]["user_id"]
            if AdminsSerializer.admin_check_permission(admin_id, 'BillManaging'):
                filters = {
                    "bill_id": bill_id,
                    "bill_serial": bill_serial,
                }
                filters = {k: v for k, v in filters.items() if v is not None}
                try:
                    bill = Bills.objects.get(**filters)
                    bill_result = bill.as_dict()
                    return True, bill_result
                except:
                    wrong_data_result["farsi_message"] = "ای دی های ورودی اشتباه است"
                    wrong_data_result["english_message"] = "input IDs are wrong"
                    return False, wrong_data_result
            else:
                return False, wrong_token_result
        else:
            return False, wrong_token_result

    @staticmethod
    def admin_edit_bill_serializer(
            token, bill_id, payment_dead_line, other_information):
        """
            param : [token, bill_id, payment_dead_line, other_information]

            return :
            A tuple containing a boolean indicating the success or failure of the operation, and a list of
            serialized data results.  it returns a false status along with an error message.
        """
        token_result = token_to_user_id(token)
        if token_result["status"] == "OK":
            admin_id = token_result["data"]["user_id"]
            if AdminsSerializer.admin_check_permission(admin_id, 'BillManaging'):
                try:
                    bill = Bills.objects.filter(bill_id=bill_id)
                    bill.update(payment_dead_line=payment_dead_line, other_information=other_information)
                except:
                    wrong_data_result["farsi_message"] = "bill_id اشتباه است"
                    wrong_data_result["english_message"] = "Wrong bill_id"
                    return False, wrong_data_result
                return True, status_success_result
        else:
            return False, wrong_token_result

    @staticmethod
    def admin_remove_bill_serializer(token, bill_id):
        """
            param : [token, bill_id]

            return :
            A tuple containing a boolean indicating the success or failure of the operation, and a list of
            serialized data results.  it returns a false status along with an error message.
        """
        token_result = token_to_user_id(token)
        if token_result["status"] == "OK":
            admin_id = token_result["data"]["user_id"]
            if AdminsSerializer.admin_check_permission(admin_id, 'BillManaging'):
                try:
                    bill = Bills.objects.get(bill_id=bill_id)
                    bill.delete()
                    return True, status_success_result
                except:
                    wrong_data_result["farsi_message"] = "bill_id اشتباه است"
                    wrong_data_result["english_message"] = "Wrong bill_id"
                    return False, wrong_data_result
            else:
                return False, wrong_token_result
        else:
            return False, wrong_token_result

    @staticmethod
    def admin_create_new_bill_serializer(
            token, water_meter_serial, payment_dead_line, other_information):
        """
            param : [token, water_meter_serial, payment_dead_line, other_information]

            return :
            A tuple containing a boolean indicating the success or failure of the operation, and a list of
            serialized data results.  it returns a false status along with an error message.
        """
        token_result = token_to_user_id(token)
        if token_result["status"] == "OK":
            admin_id = token_result["data"]["user_id"]
            if AdminsSerializer.admin_check_permission(admin_id, 'BillManaging'):
                admin = Admins.objects.get(admin_id=admin_id)

                try:
                    water_meter = WaterMeters.objects.get(water_meter_serial=water_meter_serial)
                except:
                    wrong_data_result["farsi_message"] = "water_meter_serial اشتباه وارد شده "
                    wrong_data_result["english_message"] = "water_meter_serial is wrong"
                    return False, wrong_data_result
                if 'billInfo' not in water_meter.other_information or len(
                        water_meter.other_information['billInfo']) == 0:
                    """
                    in this condition check for pattern in meter project if not exist return error 
                    """
                    # get project pattern
                    water_meter_project_obj = water_meter.water_meter_project
                    try:
                        pattern_obj = Pattern.objects.get(pattern_project=water_meter_project_obj)
                    except:
                        wrong_data_result["farsi_message"] = "رکوردهای محاسبه قیمت برای قبض اضافه نشده است"
                        wrong_data_result[
                            "english_message"] = "Price calculation records have not been added for the bill"
                        wrong_data_result["code"] = 404
                        return False, wrong_data_result
                    bill_calculation_units = pattern_obj.pattern_list
                else:
                    bill_calculation_units = water_meter.other_information['billInfo']
                bill_water_meter = water_meter
                bill_user = water_meter.water_meter_user
                bill_admin = admin
                bill_serial = random.randint(1000000000, 9999999999)
                water_meter_consumptions = WaterMetersConsumptions.objects.filter(water_meters=water_meter)
                water_meter_consumptions_bill_not_create = water_meter_consumptions.filter(bill_created=False)
                water_meter_consumptions_bill_not_create = water_meter_consumptions
                if len(water_meter_consumptions) == 0:
                    wrong_data_result["farsi_message"] = "کنتور هیچ مقدار مصرفی ندارد !!!"
                    wrong_data_result["english_message"] = "The meter has no consumption!!!"
                    wrong_data_result['code'] = 444
                    return False, wrong_data_result
                if len(water_meter_consumptions_bill_not_create) == 0:
                    wrong_data_result["farsi_message"] = "برای تمامی مقادیر مصرف قبض صادر شده"
                    wrong_data_result["english_message"] = "bill created for all consumptions"
                    wrong_data_result['code'] = 444
                    return False, wrong_data_result
                water_meter_consumptions_sorted = water_meter_consumptions_bill_not_create.order_by(
                    'create_time__date').values(
                    'create_time', 'from_previous_record', 'to_current_record')
                # print(water_meter_consumptions_sorted.first()['from_previous_record'])
                if water_meter_consumptions_sorted.first()['from_previous_record'] is not None:
                    first_consumption_record_time = water_meter_consumptions_sorted.first()['from_previous_record']
                else:
                    first_consumption_record_time = water_meter_consumptions_sorted.first()['create_time']

                bill_start_date = first_consumption_record_time
                bill_end_date = water_meter_consumptions_sorted.last()['to_current_record']
                sum_all_consumptions_value = water_meter_consumptions_bill_not_create.aggregate(Sum('value'))
                sum_all_consumptions_value = float('{:.2f}'.format(sum_all_consumptions_value['value__sum']))
                # bill_calculation_units = water_meter.other_information['billInfo']
                list_item_changed = []
                for calculation_units in bill_calculation_units:
                    changed_dict = {
                        'k': float(calculation_units['k']),
                        'v1': float(calculation_units['v1']),
                        'v2': float(calculation_units['v2'])
                    }
                    list_item_changed.append(changed_dict)
                bill_calculation_units_sort = sorted(list_item_changed, key=lambda d: d['v1'])
                bill_calculation_units_list = []
                bill_price = 0.0
                for unit in bill_calculation_units_sort:
                    if sum_all_consumptions_value > unit['v2']:
                        bill_calculation_units_list.append((unit['k'], unit['v2']))
                    if sum_all_consumptions_value < unit['v2']:
                        last_unit = sum_all_consumptions_value - unit['v1']
                        last_unit = float('{:.2f}'.format(last_unit))
                        bill_calculation_units_list.append((unit['k'], last_unit))
                        break
                for unit_tuple in bill_calculation_units_list:
                    bill_price += unit_tuple[0] * unit_tuple[1]
                try:
                    bill_object = Bills()
                    bill_object.bill_water_meter = bill_water_meter
                    bill_object.bill_user = bill_user
                    bill_object.bill_admin = bill_admin
                    bill_object.bill_serial = bill_serial
                    bill_object.bill_start_date = bill_start_date
                    bill_object.bill_end_date = bill_end_date
                    bill_object.bill_price = bill_price
                    bill_object.payment_dead_line = payment_dead_line
                    bill_object.other_information = other_information
                    bill_object.payment_dead_line = payment_dead_line
                    bill_object.consumptions = sum_all_consumptions_value
                    bill_object.save()
                    water_meter_consumptions_bill_not_create.update(bill_created=True)
                except:
                    wrong_data_result["farsi_message"] = "قبض صادر نشد !!!"
                    wrong_data_result["english_message"] = "bill dose not created !!!"
                    return False, wrong_data_result
                if water_meter.water_meter_user is not None:
                    user_phone = water_meter.water_meter_user.user_phone
                    publish_message_to_client(phone_number=user_phone, from_where='edit_device')
                    middle_admin_publish_data = {
                        'admin_phone_number': admin.admin_phone,
                        'meter_serial': water_meter.water_meter_serial,
                        'from_where': 'edit_device'
                    }
                    publish_message_to_client(publish_func='middle_admin', data=middle_admin_publish_data)

                return True, status_success_result
            else:
                return False, wrong_token_result

    @staticmethod
    def user_get_one_bill_serializer(token, bill_serial):
        """
            param : [token, bill_seria]

            return :
            A tuple containing a boolean indicating the success or failure of the operation, and a list of
            serialized data results.  it returns a false status along with an error message.
        """
        token_result = token_to_user_id(token)
        if token_result["status"] == "OK":
            user_id = token_result["data"]["user_id"]
            if user_id == user_id:
                try:
                    bill = Bills.objects.get(bill_serial=bill_serial)
                    bill_result = bill.as_dict()
                    bill_result.pop('bill_user')
                    bill_result.pop('bill_admin')
                    return True, bill_result
                except:
                    wrong_data_result["farsi_message"] = "bill_serial اشتباه است"
                    wrong_data_result["english_message"] = "Wrong bill_serial"
                    return False, wrong_data_result
            else:
                return False, wrong_token_result
        else:
            return False, wrong_token_result

    @staticmethod
    def user_get_all_bills_serializer(token, page, count, bill_start_date, bill_end_date, water_meter_serial):
        """
            param : [token, page, count, bill_start_date, bill_end_date, water_meter_serial]

            return :
            A tuple containing a boolean indicating the success or failure of the operation, and a list of
            serialized data results.  it returns a false status along with an error message.
        """

        token_result = token_to_user_id(token)
        if token_result["status"] == "OK":
            user_id = token_result["data"]["user_id"]
            fields = {
                'page': (page, int),
                'count': (count, int),
            }
            result = wrong_result(fields)
            if result == None:
                offset = int((page - 1) * count)
                limit = int(count)
                filters = {
                    'bill_user': user_id,
                    'bill_start_date__contains': bill_start_date,
                    'bill_end_date__contains': bill_end_date,
                    'bill_water_meter__water_meter_serial': water_meter_serial
                }
                filters = {k: v for k, v in filters.items() if v is not None}
                all_bills = Bills.objects.filter(**filters)
                all_bills_pagination = all_bills.order_by(
                    '-bill_start_date')[offset:offset + limit]
                bills_count = all_bills.count()
                All_bills = []
                for bill in all_bills_pagination:
                    bill = bill.as_dict()
                    if 'bill_user' in bill:
                        bill.pop('bill_user')
                    if 'bill_admin' in bill:
                        bill.pop('bill_admin')
                    bill.update({
                        "bills_count": bills_count
                    })
                    All_bills.append(bill)

                return True, All_bills
            else:
                return result
        else:
            return False, wrong_token_result

    @staticmethod
    def user_create_link_serializer(token, bill_id):
        """
            param : [token, bill_id]

            return :
            A tuple containing a boolean indicating the success or failure of the operation, and a list of
            serialized data results.  it returns a false status along with an error message.
        """
        token_result = token_to_user_id(token)
        if token_result["status"] == "OK":
            user_id = token_result["data"]["user_id"]
            # get user bill_id's
            user_bill_ids = Bills.objects.filter(bill_user=user_id).values('bill_id')
            user_bill_id_in_str = []
            for billId in user_bill_ids:
                user_bill_id_in_str.append(str(billId['bill_id']))
            if bill_id not in user_bill_id_in_str:
                wrong_data_result["farsi_message"] = "قبض جزو قبض های کاربر نمی‌باشد یا ای دی قبض اشتباه است"
                wrong_data_result["english_message"] = "bill is not among the user's bills or bill_id is wrong."
                wrong_data_result["code"] = 444
                return False, wrong_data_result
            # create link structure
            S = 50  # number of characters in the string.
            # call random.choices() string module to find the string in Uppercase + numeric data.
            ran_character = ''.join(random.choices(string.ascii_uppercase + string.digits, k=S))
            bill_link = 'http://217.144.106.32:6565/Templates/Bill/' + f'{bill_id}' + f'/{ran_character}'
            # add bill_link to dataBase and edit bill_validation too True .
            Bills.objects.filter(bill_id=bill_id).update(bill_link=bill_link, bill_link_validation=True)
            result = {
                "Message": "bill link created successfully.",
                "bill_link": bill_link
            }
            return True, result
        else:
            return False, wrong_token_result

    @staticmethod
    def user_get_link_serializer(bill_id):
        """
            param : [bill_id]

            return :
            A tuple containing a boolean indicating the success or failure of the operation, and a list of
            serialized data results.  it returns a false status along with an error message.
        """
        bill_obj = Bills.objects.get(bill_id=bill_id)
        if bill_obj.bill_link is None:
            wrong_data_result["farsi_message"] = "لینک برای قبض ساخته نشده"
            wrong_data_result["english_message"] = "The link for the bill has not been created."
            wrong_data_result["code"] = 444
            return False, wrong_data_result
        if bill_obj.bill_link_validation is False:
            wrong_data_result["farsi_message"] = "لینک منقضی شده است"
            wrong_data_result["english_message"] = "The link has expired."
            wrong_data_result["code"] = 444
            return False, wrong_data_result
        # get bill_obj data .
        bill_result = bill_obj.as_dict()
        # bill_result.pop('bill_user_info')
        bill_result.pop('bill_admin_info')
        # edit bill validation too False
        bill_obj.bill_link_validation = False
        bill_obj.save()
        return True, bill_result

    @staticmethod
    def admin_create_exel_serializer(token, project_id, base_price, all_meter_serials, meter_tag_id,
                                     start_persian_time,
                                     end_persian_time):
        """
            param : [token, project_id, base_price, all_meter_serials, meter_tag_id,
                                     start_persian_time,
                                     end_persian_time]

            return :
            A tuple containing a boolean indicating the success or failure of the operation, and a list of
            serialized data results.  it returns a false status along with an error message.
        """
        token_result = token_to_user_id(token)
        if token_result["status"] == "OK":
            admin_id = token_result["data"]["user_id"]
            if AdminsSerializer.admin_check_permission(admin_id, 'BillManaging'):
                # get_tag
                try:
                    tag_obj = WaterMetersTags.objects.get(water_meter_tag_id=meter_tag_id)
                except:
                    wrong_data_result["farsi_message"] = "meter_tag_id شتباه است"
                    wrong_data_result["english_message"] = "meter_tag_id is wrong ."
                    wrong_data_result["code"] = 444
                    return False, wrong_data_result
                tag_name = tag_obj.water_meter_tag_name
                # end get_tag

                # get project meters .
                try:
                    meter_objects = WaterMeters.objects.filter(water_meter_project__water_meter_project_id=project_id)
                    project_name = meter_objects.values('water_meter_project__water_meter_project_name')[0][
                        'water_meter_project__water_meter_project_name']
                    # meter_serials = meter_objects.values('water_meter_serial')
                    # meter_serials = list(map(lambda itm: itm['water_meter_serial'], meter_serials))
                except:
                    wrong_data_result["farsi_message"] = "project_id شتباه است"
                    wrong_data_result["english_message"] = "project_id is wrong ."
                    wrong_data_result["code"] = 444
                    return False, wrong_data_result
                # end project meters .

                # check serial is input or not
                if all_meter_serials is None or len(all_meter_serials) == 0:
                    # get all meters object with project id and tag id
                    all_meters = meter_objects.filter(
                        water_meter_type__water_meter_tag__water_meter_tag_id=meter_tag_id)
                else:
                    all_meters = meter_objects.filter(
                        water_meter_type__water_meter_tag__water_meter_tag_id=meter_tag_id)
                    meter_serials = all_meters.values('water_meter_serial')
                    meter_serials = list(map(lambda itm: itm['water_meter_serial'], meter_serials))
                    meter_not_valid = list(filter(lambda a: a not in meter_serials, all_meter_serials))
                    if len(meter_not_valid) > 0:
                        wrong_data_result["farsi_message"] = f"کنتور با شماره سریال  {meter_not_valid[0]} یافت نشد ."
                        wrong_data_result[
                            "english_message"] = f"Meter with {meter_not_valid[0]} serial number not found."
                        wrong_data_result["code"] = 444
                        return False, wrong_data_result
                # end check serial is input or not .

                # get excel template root
                base_dir = os.getcwd()
                bill_excel_folder = os.path.join(base_dir, 'bill_excels')
                # end get excel template root .

                # find structure of each category
                category_structure_path = os.path.exists(os.path.join(bill_excel_folder, f'{tag_name}.xlsx'))
                if not category_structure_path:
                    wrong_data_result["farsi_message"] = "ساختاری برای دستبندی مورد نظر یافت نشد ."
                    wrong_data_result["english_message"] = "No structure was found for the desired tag."
                    wrong_data_result["code"] = 444
                    return False, wrong_data_result
                category_structure_path = os.path.join(bill_excel_folder, f'{tag_name}.xlsx')
                # start time
                start_consumption_jalali_time = start_persian_time.split('-')
                start_consumption_jalali_time = list(map(lambda x: int(x), start_consumption_jalali_time))
                # Jalali to Gregorian
                start_time_gregorian = JalaliDate(start_consumption_jalali_time[0], start_consumption_jalali_time[1],
                                                  start_consumption_jalali_time[2]).to_gregorian()
                # end time
                end_consumption_jalali_time = end_persian_time.split('-')
                end_consumption_jalali_time = list(map(lambda x: int(x), end_consumption_jalali_time))
                # Jalali to Gregorian
                end_time_gregorian = JalaliDate(end_consumption_jalali_time[0], end_consumption_jalali_time[1],
                                                end_consumption_jalali_time[2]).to_gregorian()
                difference_days = end_time_gregorian - start_time_gregorian
                difference_days = difference_days.days + 1

                for water_meter_obj in all_meters:
                    try:
                        water_meter_serial = water_meter_obj.water_meter_serial
                        water_meter_name = water_meter_obj.water_meter_name
                        water_meter_name = water_meter_name.split('-')
                        del water_meter_name[0]
                        water_meter_name = ' '.join(map(str, water_meter_name))
                    except:
                        water_meter_serial = None
                        water_meter_name = None

                    # get sum value with filter on time
                    consumptions = WaterMetersConsumptions.objects.filter(
                        water_meters__water_meter_serial=water_meter_serial,
                        create_time__gte=start_time_gregorian, create_time__lte=end_time_gregorian)
                    if len(consumptions) > 0:
                        sum_value = consumptions.aggregate(Sum('value'))

                        sum_value = float('{:.2f}'.format(sum_value['value__sum']))
                        sum_value_cubic_meters = sum_value / 1000
                        sum_value_cubic_meters = float('{:.2f}'.format(sum_value_cubic_meters))
                        # average = sum_value / difference_days
                        # average = float('{:.2f}'.format(average))
                        average_cubic_meters = sum_value_cubic_meters / difference_days
                        average_cubic_meters = float('{:.2f}'.format(average_cubic_meters))
                        bill_serial = random.randint(1000000000, 9999999999)
                        payment_code = random.randint(100000, 999999)
                        # calculation price
                        price = int(sum_value_cubic_meters * base_price)
                        file_path = os.path.join(bill_excel_folder, category_structure_path)
                        workbook = load_workbook(filename=file_path)
                        sheet = workbook.active
                        cell_obj = sheet['Q10'].value
                        final_price = price
                        if cell_obj is not None:
                            final_price = price + cell_obj
                        price_words = digits.to_word(final_price) + ' ' + 'ریال'
                        # end calculation price .

                        # Start by opening the spreadsheet and selecting the main sheet
                        file_path = os.path.join(bill_excel_folder, category_structure_path)
                        workbook = load_workbook(filename=file_path)
                        sheet = workbook.active
                        sheet["D18"] = sum_value_cubic_meters
                        sheet["C5"] = water_meter_name
                        sheet["B13"] = water_meter_serial
                        sheet["F18"] = average_cubic_meters
                        sheet["B21"] = price_words
                        sheet["I21"] = bill_serial
                        sheet["Q21"] = final_price
                        sheet["K21"] = payment_code
                        sheet["Q9"] = price
                        project_folder = os.path.exists(os.path.join(bill_excel_folder, f'{project_name}'))
                        if not project_folder:
                            os.mkdir(os.path.join(bill_excel_folder, f'{project_name}'))
                        project_folder_path = os.path.join(bill_excel_folder, f'{project_name}')
                        tag_folder = os.path.exists(os.path.join(bill_excel_folder, project_folder_path, tag_name))
                        if not tag_folder:
                            os.mkdir(os.path.join(project_folder_path, tag_name))
                        tag_path = os.path.join(project_folder_path, tag_name)
                        final_file_path = os.path.join(tag_path, f"{water_meter_serial}.xlsx")
                        # Save the spreadsheet
                        workbook.save(filename=final_file_path)

                    else:
                        # Start by opening the spreadsheet and selecting the main sheet
                        file_path = os.path.join(bill_excel_folder, category_structure_path)
                        workbook = load_workbook(filename=file_path)
                        sheet = workbook.active
                        sheet["D18"] = '-'
                        sheet["C5"] = water_meter_name
                        sheet["B13"] = water_meter_serial
                        sheet["F18"] = '-'
                        sheet["B21"] = '-'
                        sheet["I21"] = '-'
                        sheet["Q21"] = '-'
                        sheet["K21"] = '-'
                        sheet["Q9"] = '-'
                        project_folder = os.path.exists(os.path.join(bill_excel_folder, f'{project_name}'))
                        if not project_folder:
                            os.mkdir(os.path.join(bill_excel_folder, f'{project_name}'))
                        project_folder_path = os.path.join(bill_excel_folder, f'{project_name}')
                        tag_folder = os.path.exists(os.path.join(bill_excel_folder, project_folder_path, tag_name))
                        if not tag_folder:
                            os.mkdir(os.path.join(project_folder_path, tag_name))
                        tag_path = os.path.join(project_folder_path, tag_name)
                        final_file_path = os.path.join(tag_path, f"{water_meter_serial}.xlsx")
                        # Save the spreadsheet
                        workbook.save(filename=final_file_path)
                status_success_result['code'] = 200
                return True, status_success_result
            else:
                wrong_token_result['code'] = 409
                return False, wrong_token_result
        else:
            wrong_token_result['code'] = 409
            return False, wrong_token_result

    @staticmethod
    def admin_create_power_exel_serializer(token, project_id, tag_id):
        """
            param : [token, project_id, tag_id]

            return :
            A tuple containing a boolean indicating the success or failure of the operation, and a list of
            serialized data results.  it returns a false status along with an error message.
        """
        token_result = token_to_user_id(token)
        if token_result["status"] == "OK":
            admin_id = token_result["data"]["user_id"]
            if AdminsSerializer.admin_check_permission(admin_id, 'BillManaging'):
                # get project details
                try:
                    project_obj = WaterMetersProjects.objects.get(water_meter_project_id=project_id)
                    project_name = project_obj.water_meter_project_name
                except:
                    wrong_data_result["farsi_message"] = "project_id شتباه است"
                    wrong_data_result["english_message"] = "project_id is wrong ."
                    wrong_data_result["code"] = 444
                    return False, wrong_data_result

                # end get project details.

                # get tag object detail .

                try:
                    tag_obj = WaterMetersTags.objects.get(water_meter_tag_id=tag_id)
                    tag_name = tag_obj.water_meter_tag_name
                except:
                    wrong_data_result["farsi_message"] = "project_id شتباه است"
                    wrong_data_result["english_message"] = "project_id is wrong ."
                    wrong_data_result["code"] = 444
                    return False, wrong_data_result

                # end get tag object detail .

                # read data from excel file .
                root_path = os.getcwd()
                bill_excels = os.path.join(root_path, 'bill_excels')
                detail_excel_path = os.path.join(bill_excels, 'PowerMetersDetails.xlsx')
                category_structure_path = os.path.join(bill_excels, 'برق.xlsx')
                # creating or loading an excel workbook
                newWorkbook = openpyxl.load_workbook(detail_excel_path)

                # Accessing specific sheet of the workbook.
                active_worksheet = newWorkbook["Sheet1"]

                # Passing the column index to the worksheet and traversing through the each row of the column

                for i in active_worksheet:
                    # Printing the column values of every row
                    serial = i[0].value
                    if serial is not None:
                        serial = serial.split(' ')[1]
                    meter_name = f"{i[2].value}"
                    if type(i[4].value) is float:
                        meter_value = float('{:.2f}'.format(i[4].value / 1000))
                        average = float('{:.2f}'.format(meter_value / 31))
                        price = meter_value * 310000
                        final_price = price + 1000000
                        persian_word_price = digits.to_word(int(final_price)) + ' ' + 'ریال'
                    else:
                        meter_value = None
                        average = None
                        price = None
                        final_price = None
                        persian_word_price = None

                    # Start by opening the spreadsheet and selecting the main sheet
                    file_path = os.path.join(bill_excels, category_structure_path)
                    workbook = load_workbook(filename=file_path)
                    sheet = workbook.active
                    sheet["D18"] = meter_value
                    sheet["C5"] = meter_name
                    sheet["B13"] = serial
                    sheet["F18"] = average
                    sheet["B21"] = persian_word_price
                    sheet["Q21"] = final_price
                    sheet["Q9"] = price
                    sheet["Q21"] = final_price
                    project_folder = os.path.exists(os.path.join(bill_excels, f'{project_name}'))
                    if not project_folder:
                        os.mkdir(os.path.join(bill_excels, f'{project_name}'))
                    project_folder_path = os.path.join(bill_excels, f'{project_name}')
                    tag_folder = os.path.exists(os.path.join(bill_excels, project_folder_path, tag_name))
                    if not tag_folder:
                        os.mkdir(os.path.join(project_folder_path, tag_name))
                    tag_path = os.path.join(project_folder_path, tag_name)
                    final_file_path = os.path.join(tag_path, f"{serial}.xlsx")
                    # Save the spreadsheet
                    workbook.save(filename=final_file_path)

                # end read data from excel file .

                status_success_result['code'] = 200
                return True, status_success_result
            else:
                wrong_token_result['code'] = 409
                return False, wrong_token_result
        else:
            wrong_token_result['code'] = 409
            return False, wrong_token_result

    @staticmethod
    def admin_create_new_bill_v1_serializer(
            token, water_meter_serial, payment_dead_line, other_information, start_time, end_time):
        """
            param : [token, water_meter_serial, payment_dead_line, other_information, start_time, end_time]

            return :
            A tuple containing a boolean indicating the success or failure of the operation, and a list of
            serialized data results.  it returns a false status along with an error message.
        """
        token_result = token_to_user_id(token)
        if token_result["status"] == "OK":
            admin_id = token_result["data"]["user_id"]
            if AdminsSerializer.admin_check_permission(admin_id, 'BillManaging'):
                admin = Admins.objects.get(admin_id=admin_id)

                try:
                    water_meter = WaterMeters.objects.get(water_meter_serial=water_meter_serial)
                except:
                    wrong_data_result["farsi_message"] = "water_meter_serial اشتباه وارد شده "
                    wrong_data_result["english_message"] = "water_meter_serial is wrong"
                    return False, wrong_data_result
                test_structure = {"billInfo": [
                    {
                        "k": "200000",
                        "v1": "1",
                        "v2": "1000000000"
                    }
                ]}
                bill_calculation_units = test_structure['billInfo']
                bill_water_meter = water_meter
                bill_user = water_meter.water_meter_user
                bill_admin = admin
                bill_serial = random.randint(1000000000, 9999999999)
                water_meter_consumptions = WaterMetersConsumptions.objects.filter(water_meters=water_meter,
                                                                                  create_time__gte=start_time,
                                                                                  create_time__lte=end_time)
                water_meter_consumptions_bill_not_create = water_meter_consumptions
                if len(water_meter_consumptions) == 0:
                    wrong_data_result["farsi_message"] = "کنتور هیچ مقدار مصرفی ندارد !!!"
                    wrong_data_result["english_message"] = "The meter has no consumption!!!"
                    wrong_data_result['code'] = 444
                    return False, wrong_data_result

                sum_all_consumptions_value = water_meter_consumptions_bill_not_create.aggregate(Sum('value'))
                sum_all_consumptions_value = float('{:.2f}'.format(sum_all_consumptions_value['value__sum']))
                # bill_calculation_units = water_meter.other_information['billInfo']
                list_item_changed = []
                for calculation_units in bill_calculation_units:
                    changed_dict = {
                        'k': float(calculation_units['k']),
                        'v1': float(calculation_units['v1']),
                        'v2': float(calculation_units['v2'])
                    }
                    list_item_changed.append(changed_dict)
                bill_calculation_units_sort = sorted(list_item_changed, key=lambda d: d['v1'])
                bill_calculation_units_list = []
                bill_price = 0.0
                for unit in bill_calculation_units_sort:
                    if sum_all_consumptions_value > unit['v2']:
                        bill_calculation_units_list.append((unit['k'], unit['v2']))
                    if sum_all_consumptions_value < unit['v2']:
                        last_unit = sum_all_consumptions_value - unit['v1']
                        last_unit = float('{:.2f}'.format(last_unit))
                        bill_calculation_units_list.append((unit['k'], last_unit))
                        break
                for unit_tuple in bill_calculation_units_list:
                    bill_price += unit_tuple[0] * unit_tuple[1]
                # difference between dates in timedelta
                start = datetime.strptime(start_time, "%Y-%m-%d %H:%M:%S")
                end = datetime.strptime(end_time, "%Y-%m-%d %H:%M:%S")

                days_difference = end - start
                print(days_difference)
                other_information['days_number'] = days_difference.days
                price_words = digits.to_word(int(bill_price))
                other_information['price_words'] = price_words
                try:
                    bill_object = Bills()
                    bill_object.bill_water_meter = bill_water_meter
                    bill_object.bill_user = bill_user
                    bill_object.bill_admin = bill_admin
                    bill_object.bill_serial = bill_serial
                    bill_object.bill_start_date = start_time
                    bill_object.bill_end_date = end_time
                    bill_object.bill_price = int(bill_price)
                    bill_object.payment_dead_line = payment_dead_line
                    bill_object.other_information = other_information
                    bill_object.payment_dead_line = payment_dead_line
                    bill_object.consumptions = sum_all_consumptions_value
                    bill_object.save()
                    # water_meter_consumptions_bill_not_create.update(bill_created=True)
                except:
                    wrong_data_result["farsi_message"] = "قبض صادر نشد !!!"
                    wrong_data_result["english_message"] = "bill dose not created !!!"
                    return False, wrong_data_result

                return True, status_success_result
            else:
                return False, wrong_token_result
        else:
            wrong_token_result['code'] = 403
            return False, wrong_token_result

    @staticmethod
    def admin_create_new_bill_list_serializer(
            token, meter_serial_list, payment_dead_line, other_information, start_time, end_time,
            calculate_method, project_id, tag_id):
        """
            param : [token, meter_serial_list, payment_dead_line, other_information, start_time, end_time,
                        calculate_method, project_id,]

            return :
            A tuple containing a boolean indicating the success or failure of the operation, and a list of
            serialized data results.  it returns a false status along with an error message.
        """
        token_result = token_to_user_id(token)
        if token_result["status"] == "OK":
            admin_id = token_result["data"]["user_id"]
            if AdminsSerializer.admin_check_permission(admin_id, 'BillManaging'):
                admin = Admins.objects.get(admin_id=admin_id)
                # check calculate_method
                valid_calculate_method = ['from_project', 'from_meter', 'project_priority', 'meter_priority']
                if calculate_method not in valid_calculate_method:
                    wrong_data_result[
                        "farsi_message"] = f"مقدار های مجاز برای calculate_method : {valid_calculate_method}"
                    wrong_data_result["english_message"] = f"valid calculate_method : {valid_calculate_method}"
                    wrong_data_result["code"] = 444
                    return False, wrong_data_result
                # end check calculate_method .

                response_dict = {}
                success_response_list = []
                unsuccessful_response_list = []
                # all project meter's.
                project_meters = WaterMeters.objects.filter(water_meter_project_id=project_id).values(
                    'water_meter_serial')
                project_meters = list(map(lambda x: x['water_meter_serial'], project_meters))

                for meter_serial in meter_serial_list:
                    if meter_serial in project_meters:
                        # price calculate structure
                        meter_object = WaterMeters.objects.get(water_meter_serial=meter_serial)
                        # get bill calculation coefficient from that save in meter other_information .
                        bill_factor = meter_object.other_information.get('bill_factor', None)
                        if bill_factor is None:
                            bill_factor = 1
                        # if calculate_method == 'from_meter' : get from meter other_information
                        if calculate_method == 'from_meter':
                            bill_structure = meter_object.other_information.get('billInfo', None)


                        # elif calculate_method == 'from_project' : get project pattern
                        elif calculate_method == 'from_project':
                            # try:
                            print("project_id : " , project_id)
                            patter_object = Pattern.objects.get(pattern_project=project_id , pattern_tag = tag_id)
                            bill_structure = patter_object.pattern_list
                            # except:
                            #     wrong_data_result[
                            #         "farsi_message"] = "هیچ الگوی مصرفی برای پروژه وجود ندارد"
                            #     wrong_data_result["english_message"] = "There is no  pattern for the project"
                            #     wrong_data_result["code"] = 444
                            #     return False, wrong_data_result

                        # elif calculate_method == 'project_priority' : first check project pattern if is null
                        # :: check meter other_information
                        elif calculate_method == 'project_priority':
                            patter_object = Pattern.objects.filter(pattern_project=project_id)
                            bill_structure = [pattern.as_dict() for pattern in patter_object][0]['pattern_list']
                            if len(patter_object) == 0:
                                meter_object = WaterMeters.objects.get(water_meter_serial=meter_serial)
                                bill_structure = meter_object.other_information.get('billInfo', None)
                        # elif calculate_method == 'meter_priority' : first check meter other_information if is null
                        # :: check project pattern
                        elif calculate_method == 'meter_priority':
                            meter_object = WaterMeters.objects.get(water_meter_serial=meter_serial)
                            bill_structure = meter_object.other_information.get('billInfo', None)
                            if bill_structure is None:
                                patter_object = Pattern.objects.filter(pattern_project=project_id)
                                bill_structure = [pattern.as_dict() for pattern in patter_object][0]['pattern_list']
                        # end price calculate structure .
                        if bill_structure is not None:
                            # create bill object
                            bill_calculation_units = bill_structure
                            bill_water_meter = meter_object
                            bill_user = meter_object.water_meter_user
                            bill_admin = admin
                            bill_serial = random.randint(1000000000, 9999999999)
                            water_meter_consumptions = WaterMetersConsumptions.objects.filter(water_meters=meter_object,
                                                                                              create_time__gte=start_time,
                                                                                              create_time__lte=end_time)
                            try:
                                start_period_sum = WaterMetersConsumptions.objects.filter(
                                    water_meters=meter_object,
                                    create_time__lte=start_time).aggregate(Sum('value'))['value__sum']
                                start_period_sum = float('{:.2f}'.format(start_period_sum))
                            except:
                                start_period_sum = 0

                            try:
                                end_period_sum = WaterMetersConsumptions.objects.filter(
                                    water_meters=meter_object, create_time__lte=end_time).aggregate(Sum('value'))[
                                    'value__sum']
                                end_period_sum = float('{:.2f}'.format(end_period_sum))
                            except:
                                end_period_sum = 0

                            # except:
                            #     end_period_sum = None
                            if len(water_meter_consumptions) > 0:
                                water_meter_consumptions_bill_not_create = water_meter_consumptions.filter(
                                    bill_created=False)
                                if len(water_meter_consumptions_bill_not_create) > 0:
                                    sum_all_consumptions_value = water_meter_consumptions_bill_not_create.aggregate(
                                        Sum('value'))
                                    sum_all_consumptions_value = float(
                                        '{:.2f}'.format(sum_all_consumptions_value['value__sum']))
                                    # bill_calculation_units = water_meter.other_information['billInfo']
                                    list_item_changed = []
                                    for calculation_units in bill_calculation_units:
                                        changed_dict = {
                                            'k': float(calculation_units['k']),
                                            'v1': float(calculation_units['v1']),
                                            'v2': float(calculation_units['v2'])
                                        }
                                        list_item_changed.append(changed_dict)
                                    bill_calculation_units_sort = sorted(list_item_changed, key=lambda d: d['v1'])
                                    bill_calculation_units_list = []
                                    bill_price = 0.0
                                    for unit in bill_calculation_units_sort:
                                        if sum_all_consumptions_value > unit['v2']:
                                            bill_calculation_units_list.append((unit['k'], unit['v2']))
                                        if sum_all_consumptions_value < unit['v2']:
                                            last_unit = sum_all_consumptions_value - unit['v1']
                                            last_unit = float('{:.2f}'.format(last_unit))
                                            bill_calculation_units_list.append((unit['k'], last_unit))
                                            break
                                    for unit_tuple in bill_calculation_units_list:
                                        bill_price += unit_tuple[0] * unit_tuple[1]
                                    # difference between dates in timedelta
                                    start = datetime.strptime(start_time, "%Y-%m-%d")
                                    end = datetime.strptime(end_time, "%Y-%m-%d")
                                    if end_period_sum != 0 and start_period_sum != 0:
                                        consumption_checker = end_period_sum - start_period_sum
                                        if consumption_checker != sum_all_consumptions_value:
                                            if consumption_checker > sum_all_consumptions_value:
                                                consumption_difference = consumption_checker - sum_all_consumptions_value
                                                start_period_sum = start_period_sum - consumption_difference
                                            elif consumption_checker < sum_all_consumptions_value:
                                                consumption_difference = sum_all_consumptions_value - consumption_checker
                                                start_period_sum = start_period_sum - consumption_difference
                                    days_difference = end - start
                                    other_information['period_days'] = days_difference.days
                                    print(f"bill_price : {bill_price} - bill_factor : {bill_factor}")
                                    final_price = int(float(bill_price) * float(bill_factor))
                                    price_words = digits.to_word(final_price)
                                    other_information['price_words'] = price_words
                                    other_information['start_period_sum'] = start_period_sum / 1000
                                    other_information['end_period_sum'] = end_period_sum / 1000
                                    other_information['bill_factor'] = bill_factor
                                    try:
                                        bill_object = Bills()
                                        bill_object.bill_water_meter = bill_water_meter
                                        bill_object.bill_user = bill_user
                                        bill_object.bill_admin = bill_admin
                                        bill_object.bill_serial = bill_serial
                                        bill_object.bill_start_date = start_time
                                        bill_object.bill_end_date = end_time
                                        bill_object.bill_price = final_price / 1000
                                        bill_object.payment_dead_line = payment_dead_line
                                        bill_object.other_information = other_information
                                        bill_object.payment_dead_line = payment_dead_line
                                        bill_object.consumptions = sum_all_consumptions_value / 1000
                                        bill_object.save()
                                        response_dict = {
                                            "meter_serial": f'{meter_serial}',
                                            "farsi_message": "با موفقیت انجام شد.",
                                            "english_message": "Successfully Done.",
                                        }
                                        success_response_list.append(response_dict)
                                        water_meter_consumptions_bill_not_create.update(bill_created=True)
                                        user_phone = meter_object.water_meter_user.user_phone
                                        publish_message_to_client(phone_number=user_phone, from_where='bill_created')
                                        middle_admin_publish_data = {
                                            'admin_phone_number': admin.admin_phone,
                                            # 'project_id': meter_object.water_meter_project.water_meter_project_id,
                                            'meter_serial': meter_object.water_meter_serial,
                                            'from_where': 'bill_created'
                                        }
                                        publish_message_to_client(publish_func='middle_admin',
                                                                  data=middle_admin_publish_data)

                                    except:
                                        response_dict = {
                                            "meter_serial": f'{meter_serial}',
                                            "farsi_message": "قبض ایجاد نشد.",
                                            "english_message": "bill not created."
                                        }
                                        unsuccessful_response_list.append(response_dict)
                                        # end create bill object .
                                else:
                                    response_dict = {
                                        "meter_serial": f'{meter_serial}',
                                        "farsi_message": "قبض تکراری (برای بازه ی انتخابی قبلا قبض صادر شده)",
                                        "english_message": "Duplicate bill (for the selected period, "
                                                           "the bill was previously issued)"
                                    }
                                    unsuccessful_response_list.append(response_dict)

                            else:
                                response_dict = {
                                    "meter_serial": f'{meter_serial}',
                                    "farsi_message": "کنتور در بازه انتخاب شده مصرفی نداشته.",
                                    "english_message": "The meter has no consumption in the selected period."
                                }
                                unsuccessful_response_list.append(response_dict)
                        else:
                            response_dict = {
                                "meter_serial": f'{meter_serial}',
                                "farsi_message": "الگوی محاسبه قبض یافت نشد.",
                                "english_message": "bill pattern not found."
                            }
                            unsuccessful_response_list.append(response_dict)
                    else:
                        response_dict = {
                            "meter_serial": f'{meter_serial}',
                            "farsi_message": "کنتور جزو کنتور های پروژه نیست",
                            "english_message": "The meter is not among the project meters"
                        }
                        unsuccessful_response_list.append(response_dict)
                result = {
                    "all_meters_inputed": len(meter_serial_list),
                    "all_bill_successfully_created": {
                        "count": len(success_response_list),
                        "details": success_response_list
                    },
                    "all_bill_not_created": {
                        "count": len(unsuccessful_response_list),
                        "details": unsuccessful_response_list
                    }
                }
                return True, result
            else:
                wrong_token_result['code'] = 444
                wrong_token_result['farsi_message'] = "دسترسی غیرمجاز"
                return False, wrong_token_result
        else:
            wrong_token_result['code'] = 403
            return False, wrong_token_result

    @staticmethod
    def get_one_bill_serializer(bill_id, bill_serial):
        """
            param : [bill_id, bill_serial]

            return :
            A tuple containing a boolean indicating the success or failure of the operation, and a list of
            serialized data results.  it returns a false status along with an error message.
        """
        filters = {
            "bill_id": bill_id,
            "bill_serial": bill_serial,
        }
        filters = {k: v for k, v in filters.items() if v is not None}
        try:
            bill = Bills.objects.get(**filters)
            bill_result = bill.as_dict()
            return True, bill_result
        except:
            wrong_data_result["farsi_message"] = "ای دی های ورودی اشتباه است"
            wrong_data_result["english_message"] = "input IDs are wrong"
            return False, wrong_data_result
