import os
import openpyxl
import sys
import datetime
import csv
import django
from pathlib import Path
from persiantools.jdatetime import JalaliDateTime
from collections import OrderedDict
from openpyxl import Workbook

# ------------------------------------define django setting to access project model-------------------------------------
base_dir = os.getcwd()
project_path = Path(base_dir).parent
sys.path.append(f"{project_path}")
os.environ.setdefault("DJANGO_SETTINGS_MODULE", "AutomationSayalSanjesh.settings")
django.setup()
# ----------------------------------------------------------------------------------------------------------------------

# --------------------------------------------Read From CSV-------------------------------------------------------------
from django.db.models import Sum
from persiantools.jdatetime import JalaliDate
from SayalSanjesh.models import WaterMeters, WaterMetersConsumptions
from Authorization.models import StaticToken
from SayalSanjesh.Serializers.WaterMeterSerializer import WaterMeterSerializer
from Authorization.Serializers.AdminsSerializer import AdminsSerializer

serializer_object = WaterMeterSerializer()
get_all_power_meter_serial = list(WaterMeters.objects.all().values('water_meter_serial'))
get_all_power_meter_serial = list(map(lambda x: x['water_meter_serial'], get_all_power_meter_serial))
static_token = list(StaticToken.objects.all().values('token'))[0]['token']


# -----------------------------------------------ExcelManager-----------------------------------------------------------
# def all_directory_file():
#     # Replace 'path_to_your_directory' with the actual directory path
#     directory = os.getcwd()
#     directory = os.path.join(directory, 'BaseExcelFile')
#     # Get a list of all files in the directory
#     files = os.listdir(directory)
#     return files


class ExcelManager:
    # def __init__(self):
    #     self.all_file_in_directory = all_directory_file()
    #     # self.valid_file_name = ['EXC.xlsx', 'EXC04.xlsx', 'EXC05.xlsx', 'EXC07.xlsx', 'EXC08.xlsx',
    #     #                         'EXC23.xlsx', 'EXC24.xlsx', 'EXC27.xlsx']
    #     self.valid_file_name = ['EXC06.xlsx']
    #     self.new_file_names = {
    #         # 'EXC.xlsx': 'SEMK-02311101.xlsx',
    #         # 'EXC23.xlsx': 'SEMK-02311123.xlsx',
    #         # 'EXC24.xlsx': 'SEMK-02311124.xlsx',
    #         # 'EXC27.xlsx': 'SEMK-02311127.xlsx',
    #         # 'EXC04.xlsx': 'SEMK-02311104.xlsx',
    #         # 'EXC05.xlsx': 'SEMK-02311105.xlsx',
    #         'EXC06.xlsx': 'SEMK-02311106.xlsx',
    #         # 'EXC07.xlsx': 'SEMK-02311107.xlsx',
    #         # 'EXC08.xlsx': 'SEMK-02311108.xlsx',
    #     }
    #     self.file_created = []

    def read_from_excel(self, file_name, file_path):
        # Open the Excel file
        workbook = openpyxl.load_workbook(file_path)
        # Select the active worksheet
        sheet = workbook.active
        # Iterate through the rows and print the data line by line
        sum_by_date = {}
        for row in sheet.iter_rows(values_only=True):
            if type(row[0]) == datetime.datetime:
                time = row[0].date()
                # sum_day = row[1] + row[2] + row[3]
                sum_day = row[1]
                if time in sum_by_date:
                    sum_by_date[time] += sum_day
                else:
                    sum_by_date[time] = sum_day
        return sum_by_date

    def create_new_exel_file(self):
        # Get the current working directory
        base_dir = os.getcwd()

        # Specify the new directory name
        new_directory_name = 'NewExcelFile'

        # Create the full path for the new directory
        new_directory_path = os.path.join(base_dir, new_directory_name)

        # Check if the directory exists
        if not os.path.exists(new_directory_path):
            # Create a new directory if it doesn't exist
            os.mkdir(new_directory_path)
        # Write the data
        for file_name in self.all_file_in_directory:
            if file_name in self.valid_file_name and file_name not in self.file_created:
                file_path = os.path.join(base_dir, 'BaseExcelFile', file_name)
                # read from csv
                sum_by_date = self.read_from_excel(file_name=file_name, file_path=file_path)
                # Create a new workbook
                wb = Workbook()
                ws = wb.active
                # ws.title = "Sum_By_Date"
                # Write the headers
                ws['A1'] = "Time"
                ws['B1'] = "Value"
                row_num = 2
                for date, total_sum in sum_by_date.items():
                    ws.cell(row=row_num, column=1, value=date.strftime('%Y-%m-%d'))
                    ws.cell(row=row_num, column=2, value=total_sum)
                    row_num += 1
                # Save the workbook
                new_file_name = self.new_file_names[file_name]

                final_path = os.path.join(new_directory_path, new_file_name)
                wb.save(final_path)
                self.file_created.append(file_name)

    def add_to_data_base_from_excel(self):
        # Path to the folder containing Excel files
        base_dir = os.getcwd()
        folder_path = os.path.join(base_dir, 'NewExcelFile')
        # Get a list of all Excel files in the folder
        excel_files = [f for f in os.listdir(folder_path) if f.endswith('.xlsx')]

        utc_time = JalaliDate(1402, 8, 20).to_gregorian()

        # Loop through each Excel file and read line by line
        check_field = []
        for file in excel_files:
            file_name, file_extension = os.path.splitext(file)
            file_path = os.path.join(folder_path, file)
            workbook = openpyxl.load_workbook(file_path)
            sheet = workbook.active
            print(f"************{file_name}***************")
            for row in sheet.iter_rows(values_only=True):
                # try:
                if row[0] != 'date' and row[0] != 'date ':
                    if file_name in get_all_power_meter_serial:
                        meter_serial = file_name
                        time = row[0]
                        if row[0] is not None:
                            if type(row[0]) != str:
                                date = row[0].date()
                                # dt = datetime.datetime(date.year, date.month, date.day, 19, 30, 0)
                                dt = JalaliDateTime(date.year, date.month, date.day, 19, 30, 0).to_gregorian()
                            elif type(row[0]) == str and row[0] != 'Date':
                                date = row[0].split('/')
                                # dt = datetime.datetime(int(date[0]), int(date[1]), int(date[2]), 19, 30, 0)
                                dt = JalaliDateTime(int(date[0]), int(date[1]), int(date[2]), 19, 30, 0).to_gregorian()
                            jalali_date = JalaliDateTime.to_jalali(dt)
                            value = row[1]
                            if value is not None:
                                value = int(row[1] * 1000)
                                print(f"date : {dt} -- value : {value}")
                                res = WaterMetersConsumptions.objects.create(water_meters_id=meter_serial,
                                                                             value=value,
                                                                             create_time=dt, information={},
                                                                             cumulative_value=None,
                                                                             from_previous_record=None,
                                                                             to_current_record=None)
                                with open('WaterMeterAddFromExcel.txt', 'a') as file:
                                    file.write(
                                        f"{meter_serial} - add_data : {value} - add_time {dt}\n")
                #             # valid_days = [11, 12, 13, 14, 15, 16, 17, 18, 19, 20, 21, 22, 23, 24, 25, 26, 27, 28]
                #             valid_days = [11, 12, 13, 14, 15, 16, 17, 18, 19, 20, 21, 22]
                #             if dt.month == 11 and dt.day in valid_days:
                #
                #                 print(res)

    def write_cumulative_api_to_csv(self, meter_serial_list):
        # get admin token
        admin_class = AdminsSerializer()
        login_func = admin_class.admin_login_serializer(admin_phone='09356165600', admin_password='1234')
        meter_class = WaterMeterSerializer()
        token = login_func[1].get('token')
        start_time = "2023-10-14"
        end_time = "2023-10-23"
        for serial in meter_serial_list:
            print(serial)
            # call cumulative function
            cumulative_func_response = meter_class.admin_get_all_consumptions_by_date_serializer(
                token=token, water_meters=serial, project_id=None, tag_id=None, type_id=None, start_time=start_time,
                end_time=end_time, user_id=None, page=1, count=1000)
            cumulative_data = cumulative_func_response[1]
            reversed_dict = OrderedDict(reversed(list(cumulative_data.items())))
            final_data_list = []

            for value in reversed_dict:
                if value != 'average' and value != 'total' and value != 'cumulative_consumptions' and value != '2023-10-23':
                    int_num = int(float(reversed_dict[value]))
                    # Date and value to be saved
                    date_split = value.split('-')
                    jalali_date = JalaliDate(datetime.date(int(date_split[0]), int(date_split[1]), int(date_split[2])))
                    prepared_data = {
                        "day": jalali_date,
                        "value": int_num,
                    }
                    final_data_list.append(prepared_data)
            # CSV file path
            csv_file = f"{serial}.csv"
            # Write data to CSV file
            with open(csv_file, mode='w', newline='') as file:
                writer = csv.writer(file)
                writer.writerow(["day", "value"])  # Write header
                for item in final_data_list:
                    writer.writerow([item["day"], item["value"]])  # Write each row


excel_manager_object = ExcelManager()
# excel_manager_object.create_new_exel_file()
excel_manager_object.add_to_data_base_from_excel()
# ----------------------------------------------------------------------------------------------------------------------
